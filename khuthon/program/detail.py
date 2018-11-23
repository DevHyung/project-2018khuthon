#-*-encoding:utf8-*-
'''
@desc:
    프로그램 기능정리.xlsx -> 상세내역 시트의 기능을 실행하는 프로그램
@Env:
    OS : Windows 10 64bit, macOS High Sierra 10 +
    Interpreter : python 3.X
'''
__author__ = 'HyungJune,Park'
__email__  = 'khuphj@gmail.com'


from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook,Workbook
import requests
import os
import random
import time
def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for data in _DATA:
            sheet.append(data)
        book.save(_FILENAME)
    else:  # 새로만드는건
        book = Workbook()
        sheet = book.active
        sheet.title = 'result'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 40
        sheet.column_dimensions['F'].width = 15
        book.save(_FILENAME)

def get_bs4_from_file(_filepath):
    f = open(_filepath,'r',encoding='utf8')
    bs4 =  BeautifulSoup(f.read(),'lxml')
    f.close()
    return bs4

def get_kin_info():
    dataList = []
    dataList.append(['키워드','모바일 조회수','피씨 조회수'])
    dataList.append([k, '', ''])
    dataList.append(['', '', ''])
    dataList.append(['', '지식인', ''])
    dataList.append(['통검갯수', '모통갯수', ''])
    #variable
    pcTitleList = [] # 제목리스트
    pcLinkList = [] # 질문 링크리스트
    pcRegisterDateList = []
    pcReplyCntList = []
    pcReplyGrade2dList = []
    pcLike2dList = []
    #
    bs4 = BeautifulSoup(driver.page_source,'lxml')
    #bs4 = get_bs4_from_file('html.txt') #이부분도 셀레니움으로해야함
    div = bs4.find('div',class_='kinn section _kinBase')
    ul = div.find('ul',class_='type01')
    lis = ul.find_all('li')
    pcTotalCnt = len(lis)

    dataList.append([pcTotalCnt, '@', ''])
    dataList.append(
        ['', '', '순위', '제목', '	주소', '날짜', '	답변갯수', '답변1 등급', '답변2 등급', '답변3 등급', '	답변4 등급', '답변5 등급', '좋아요',
         '유익해요', '재밌어요', '동의못해요', '광고같아요'])

    for li in lis:
        link  = li.find('dt').a['href']
        pcLinkList.append(link)

    for link in pcLinkList:
        gradeList = ['','','','','']
        gradeIdx = 0
        likeList = [0,0,0,0,0]
        driver.get(link)
        time.sleep(random.randint(2, 5) / 10)
        dbs4 = BeautifulSoup(driver.page_source,'lxml')# 이부분은 셀레니움으로 해야함
        #dbs4 = BeautifulSoup(requests.get(link).text,'lxml') #이부분은 셀레니움으로 해야함
        title = dbs4.find('div',class_='title').get_text().strip()
        registerDate = dbs4.find('span',class_='c-userinfo__date').get_text().strip()
        replyCnt = dbs4.find('em',class_='_answerCount num').get_text().strip()
        divs = dbs4.find('div',class_='answer-content__list _answerList').find_all('div',class_='answer-content__item _contentWrap _answer')
        for div in divs:
            grade = div.find('div',class_='c-userinfo').em.get_text().strip()
            try:
                likes = div.find('div',class_='c-userinfo-answer').find_all('span',class_='u_likeit_list_count _count')
            except:
                break
            for idx in range(5):
                likeList[idx] = int(likes[idx].get_text().strip()) + likeList[idx]
            gradeList[gradeIdx] = grade
            gradeIdx+= 1
        pcReplyCntList.append(replyCnt)
        pcRegisterDateList.append(registerDate.replace('작성일','')[:-1])
        pcTitleList.append(title)
        pcReplyGrade2dList.append(gradeList)
        pcLike2dList.append(likeList)


    for idx in range(len(pcTitleList)):
        dataList.append(['','',idx+1,pcTitleList[idx],pcLinkList[idx],pcRegisterDateList[idx],pcReplyCntList[idx]]+pcReplyGrade2dList[idx]+pcLike2dList[idx])
    dataList.append(['','',''])
    save_excel(FILENAME, dataList, None)





if __name__== "__main__":
    # init
    '''----------------- 환경설정 부분 ---------------------'''
    FILENAME = input(">>> 저장 엑셀 파일명만 입력하세요 :").strip()+".xlsx"
    HEADER = ['', '', '', '', '', '']
    save_excel(FILENAME,None,HEADER)
    '''____  파싱부분 ____'''
    driver = webdriver.Chrome('./chromedriver')
    while True:
        driver.get('https://search.naver.com/search.naver?ie=utf8&query=')
        k = input(">>> 키워드를 입력하세요 ::").strip()
        driver.find_element_by_xpath('//*[@id="nx_query"]').clear()
        driver.find_element_by_xpath('//*[@id="nx_query"]').send_keys(k+'\n')
        get_kin_info()


